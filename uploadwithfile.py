#!/usr/bin/env python
#
# Copyright (c) 2018  Sari Fernandez <sarifern@cisco.com>
# All rights reserved.
#
# Redistribution and use in source and binary forms, with or without
# modification, are permitted provided that the following conditions
# are met:
# 1. Redistributions of source code must retain the above copyright
#    notice, this list of conditions and the following disclaimer.
# 2. Redistributions in binary form must reproduce the above copyright
#    notice, this list of conditions and the following disclaimer in the
#    documentation and/or other materials provided with the distribution.
#
# THIS SOFTWARE IS PROVIDED BY THE AUTHOR AND CONTRIBUTORS ``AS IS'' AND
# ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE
# IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE
# ARE DISCLAIMED.  IN NO EVENT SHALL THE AUTHOR OR CONTRIBUTORS BE LIABLE
# FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL
# DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS
# OR SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION)
# HOWEVER CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT
# LIABILITY, OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY
# OUT OF THE USE OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF
# SUCH DAMAGE.
#
# This script uploads information from an XLS file to a MongoDB database

from openpyxl import load_workbook
from pymongo import MongoClient
import sys
import json

def create_mongodb(config):
    """
        Creates MongoDB Connection and returns the MongoDB client

        Required input parameters:
        -config: json object containing mongoServers (with host and port) and replicaSet(if available) 
        
        Output:
        -MongoDB client object
    """

    
    mongo_url = "mongodb://"
    mongo_url += ",".join(map(lambda srv: srv['host'] + ":" + str(srv['port']), config['data']['mongoServers']))
   
    if 'replica' in config['data']:
            mongo_url += "/?replicaSet={0}".format(config['data']['replica'])

    client = MongoClient(mongo_url)

    return client

#Load configuration variables
with open('config.json', 'r') as f:
    config = json.load(f)


#Create the DB connection

mongo_db_cnxn = create_mongodb(config)
db = mongo_db_cnxn[config['data']['database']]
result = db.authenticate(config['data']['username'], config['data']['password'])
print("The DB is online: {}".format(result))

#delete the collection to keep info up to date
db.devices.drop()
wb = load_workbook(sys.argv[1])
ws = wb.active

for row in ws.iter_rows(min_row=2, max_col=3):
    #db.devices.insert({'idDevice':ws.cell(row=row,column=0),'description':ws.cell(row=row,column=1),'listPrice':ws.cell(row=row,column=2)})
	temp_row=[]
	for cell in row:
		temp_row.append(cell.value)

	db.devices.insert({'idDevice':temp_row[0],'description':temp_row[1],'listPrice':temp_row[2]})